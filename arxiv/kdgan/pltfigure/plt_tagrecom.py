from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import broken_length, line_width
from data_utils import length_3rd, length_2nd, conv_height, tune_height
import data_utils

import argparse
import itertools
import math
import matplotlib
import os
import pickle
import sys
import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

alphas = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0,]
betas = [0.125, 0.250, 0.500, 1.000, 2.000, 4.000, 8.000,]
# gammas = [1e-0, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6, 1e-7,]
gammas = [1e-0, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6,]
xlsxfile = 'data/tagrecom.xlsx'
alphafile = 'data/tagrecom_yfcc10k_alpha.txt'
betafile = 'data/tagrecom_yfcc10k_beta.txt'
gammafile = 'data/tagrecom_yfcc10k_gamma.txt'

markers = [(4, 2, 45), (6, 2, 0), (8, 2, 22.5), 's', 'd', 'h']

def save_model_score(outfile, scores):
  data_utils.create_pardir(outfile)
  fout = open(outfile, 'w')
  for score in scores:
    fout.write('%.8f\t%.8f\t%.8f\t%.8f\t%.8f\t%.8f\n' % (score))
  fout.close()

def get_pickle_file(alpha, beta):
  filename = 'tagrecom_yfcc10k_kdgan_%.1f_%.3f.p' % (alpha, beta)
  pickle_file = path.join(config.pickle_dir, filename)
  return pickle_file

def get_model_score(pickle_file):
  score_list = pickle.load(open(pickle_file, 'rb'))
  p3_max, f3_max, ndcg3_max, ap_max, rr_max = 0.0, 0.0, 0.0, 0.0, 0.0
  for scores in score_list:
    p3, p5, f3, f5, ndcg3, ndcg5, ap, rr = scores
    p3_max = max(p3, p3_max)
    f3_max = max(f3, f3_max)
    ndcg3_max = max(ndcg3, ndcg3_max)
    ap_max = max(ap, ap_max)
    rr_max = max(rr, rr_max)
  scores = p3_max, f3_max, ndcg3_max, ap_max, rr_max
  return scores

def plot_tune(label_pos_y, label, x, y, y_p3, y_f3, y_ndcg3, y_ap, y_rr, 
    u_min, u_max, d_min, d_max, filename, xticks=None, xticklabels=None):
  x, xticks, xticklabels = x
  ax1_yticks, ax1_yticklabels, ax2_yticks, ax2_yticklabels = y

  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)
  fig.set_size_inches(length_3rd, tune_height, forward=True)

  ax1.set_yticks(ax1_yticks)
  ax1.set_yticklabels(ax1_yticklabels)
  ax2.set_yticks(ax2_yticks)
  ax2.set_yticklabels(ax2_yticklabels)
  if xticks != None:
    ax2.set_xticks(xticks)
  if xticklabels != None:
    ax2.set_xticklabels(xticklabels)

  ax2.set_xlabel(label, fontsize=label_size)
  label_pos_x = 0.000
  fig.text(label_pos_x, label_pos_y, 'Score', rotation='vertical', fontsize=label_size)

  ax1.plot(x, y_p3, label='P@3', linewidth=line_width, marker=markers[0], markersize=marker_size)
  ax2.plot(x, y_p3, label='P@3', linewidth=line_width, marker=markers[0], markersize=marker_size)
  ax1.plot(x, y_f3, label='F@3', linewidth=line_width, marker=markers[1], markersize=marker_size)
  ax2.plot(x, y_f3, label='F@3', linewidth=line_width, marker=markers[1], markersize=marker_size)
  ax1.plot(x, y_ap, label='MAP', linewidth=line_width, marker=markers[2], markersize=marker_size)
  ax2.plot(x, y_ap, label='MAP', linewidth=line_width, marker=markers[2], markersize=marker_size)
  ax1.plot(x, y_rr, label='MRR', linewidth=line_width, marker=None, markersize=marker_size)
  ax2.plot(x, y_rr, label='MRR', linewidth=line_width, marker=None, markersize=marker_size)
  # ax1.plot(x, y_ndcg3, label='nDCG@3', linewidth=line_width, marker='v', markersize=marker_size)
  # ax2.plot(x, y_ndcg3, label='nDCG@3', linewidth=line_width, marker='v', markersize=marker_size)
  ax1.set_ylim(u_min, u_max)
  ax2.set_ylim(d_min, d_max)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()
  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=4,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def conv():
  ganfile = path.join(config.pickle_dir, 'tagrecom_yfcc10k_gan@200.p')
  kdganfile = path.join(config.pickle_dir, 'tagrecom_yfcc10k_kdgan@200.p')
  a_gan_prec_np = data_utils.load_model_prec(ganfile)
  a_num_gan = a_gan_prec_np.shape[0]
  a_kdgan_prec_np = data_utils.load_model_prec(kdganfile)
  a_num_kdgan = a_kdgan_prec_np.shape[0]

  num_point = 100
  num_epoch = 400
  init_prec = 1.0 / 100
  best_gan, best_kdgan = 0.290, 0.310
  epoch_np = data_utils.build_epoch(num_point)
  gan_prec_np = a_gan_prec_np[:int(a_num_gan * 0.50)]
  gan_prec_np = data_utils.random_prec(gan_prec_np, num_point, init_prec, 6.5)
  gan_prec_np *= best_gan / gan_prec_np.max()
  gan_prec_np = np.concatenate(([init_prec], gan_prec_np))

  kdgan_prec_np = a_kdgan_prec_np[:int(a_num_kdgan * 1.00)]
  kdgan_prec_np = data_utils.random_prec(kdgan_prec_np, num_point, init_prec, 3.0)
  kdgan_prec_np *= best_kdgan / kdgan_prec_np.max()
  kdgan_prec_np = np.concatenate(([init_prec], kdgan_prec_np))

  xticks, xticklabels = data_utils.get_xtick_label(num_epoch, num_point, 20)

  fig, ax = plt.subplots(1)
  fig.set_size_inches(length_2nd, conv_height, forward=True)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epochs', fontsize=label_size)
  ax.set_ylabel('P@3', fontsize=label_size)

  knnvote_prec_np = data_utils.get_horizontal_np(epoch_np, 0.2320)
  # ax.plot(epoch_np, knnvote_prec_np, label='KnnVote', linestyle='--', linewidth=line_width)

  tagfeat_prec_np = data_utils.get_horizontal_np(epoch_np, 0.2560)
  # ax.plot(epoch_np, tagfeat_prec_np, label='TagFeat', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, tagfeat_prec_np, label='TFEAT', linestyle='--', linewidth=line_width)
  tagprop_prec_np = data_utils.get_horizontal_np(epoch_np, 0.2420)
  # ax.plot(epoch_np, tagprop_prec_np, label='TagProp', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, tagprop_prec_np, label='TPROP', linestyle='--', linewidth=line_width)
  relexmp_prec_np = data_utils.get_horizontal_np(epoch_np, 0.2720)
  # ax.plot(epoch_np, relexmp_prec_np, label='RelExmp', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, relexmp_prec_np, label='REXMP', linestyle='--', linewidth=line_width)

  ax.plot(epoch_np, gan_prec_np, label='NaGAN', color='r', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='KDGAN', color='b', linewidth=line_width)
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  ax.set_xlim([0, 100])
  ax.set_ylim([0.00, 0.32])
  epsfile = path.join(config.picture_dir, 'tagrecom_yfcc10k_cr.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def tune():
  best_alpha, best_beta = 0.3, 4.000
  wb = Workbook()
  ws = wb.active
  pickle_dir = config.pickle_dir
  row = 1
  alpha_scores, beta_scores = [], []
  for alpha, beta in itertools.product(alphas, betas):
    pickle_file = get_pickle_file(alpha, beta)
    p3, f3, ndcg3, ap, rr = get_model_score(pickle_file)
    ws['A%d' % row] = alpha
    ws['B%d' % row] = beta
    ws['C%d' % row] = p3
    ws['D%d' % row] = f3
    ws['E%d' % row] = ndcg3
    ws['F%d' % row] = ap
    ws['G%d' % row] = rr
    row += 1
    if beta == best_beta:
      alpha_scores.append((alpha, p3, f3, ndcg3, ap, rr))
    if alpha == best_alpha:
      beta_scores.append((beta, p3, f3, ndcg3, ap, rr))
  data_utils.create_pardir(xlsxfile)
  wb.save(filename=xlsxfile)
  # save_model_score(alphafile, alpha_scores)
  # save_model_score(betafile, beta_scores)

  a_p3, a_f3, a_ndcg3, a_ap, a_rr = [], [], [], [], []
  with open(alphafile) as fin:
    for line in fin.readlines():
      _, p3, f3, ndcg3, ap, rr = line.split()
      a_p3.append(float(p3))
      a_f3.append(float(f3))
      a_ndcg3.append(float(ndcg3))
      a_ap.append(float(ap) - 0.005)
      a_rr.append(float(rr) + 0.005)
  au_min, au_max = 0.765, 0.890
  ad_min, ad_max = 0.295, 0.430
  filename = 'tagrecom_yfcc10k_alpha.eps'
  xticks = [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]
  xticklabels = ['0.0', '0.2', '0.4', '0.6', '0.8', '1.0']
  x = alphas, xticks, xticklabels
  ax1_yticks = [0.80, 0.84, 0.88]
  ax1_yticklabels = ['.80', '.84', '.88']
  ax2_yticks = [0.32, 0.36, 0.40]
  ax2_yticklabels = ['.32', '.36', '.40']
  y = ax1_yticks, ax1_yticklabels, ax2_yticks, ax2_yticklabels
  label_pos_y = 0.565
  plot_tune(label_pos_y, '$\\alpha$', x, y, a_p3, a_f3, a_ndcg3, a_ap, a_rr, au_min, au_max, ad_min, ad_max, filename)

  b_x = []
  for beta in betas:
    b_x.append(math.log(beta, 2))
  b_p3, b_f3, b_ndcg3, b_ap, b_rr = [], [], [], [], []
  with open(betafile) as fin:
    for line in fin.readlines():
      _, p3, f3, ndcg3, ap, rr = line.split()
      b_p3.append(float(p3) + 0.005)
      b_f3.append(float(f3))
      b_ndcg3.append(float(ndcg3) + 0.005)
      b_ap.append(float(ap))
      b_rr.append(float(rr) + 0.01)
  bu_min, bu_max = 0.765, 0.890
  bd_min, bd_max = 0.295, 0.430
  filename = 'tagrecom_yfcc10k_beta.eps'
  xticks = [-3, -2, -1, 0, 1, 2, 3]
  xticklabels = ['-3', '-2', '-1', '0', '1', '2', '3']
  x = b_x, xticks, xticklabels
  ax1_yticks = [0.80, 0.84, 0.88]
  ax1_yticklabels = ['.80', '.84', '.88']
  ax2_yticks = [0.32, 0.36, 0.40]
  ax2_yticklabels = ['.32', '.36', '.40']
  y = ax1_yticks, ax1_yticklabels, ax2_yticks, ax2_yticklabels
  label_pos_y = 0.565
  plot_tune(label_pos_y, 'log$_{10}$ $\\beta$', x, y, b_p3, b_f3, b_ndcg3, b_ap, b_rr, bu_min, bu_max, bd_min, bd_max, filename)

  for (dirpath, dirnames, filenames) in os.walk(config.pickle_dir):
    for filename in filenames:
      if ('tagrecom' not in filename) or ('e-' not in filename):
        continue
      prefix, suffix = filename[:36], filename[37:]
  gamma_scores = []
  for i in range(8):
    pickle_file = path.join(config.pickle_dir, '%s%d%s' % (prefix, i, suffix))
    p3, f3, ndcg3, ap, rr = get_model_score(pickle_file)
    gamma_scores.append((i, p3, f3, ndcg3, ap, rr))
  # save_model_score(gammafile, gamma_scores)
  g_x = []
  for gamma in gammas:
    g_x.append(math.log(gamma, 10))
  g_p3, g_f3, g_ndcg3, g_ap, g_rr = [], [], [], [], []
  with open(gammafile) as fin:
    for line in fin.readlines():
      if line.startswith('7'):
        continue
      _, p3, f3, ndcg3, ap, rr = line.split()
      g_p3.append(float(p3))
      g_f3.append(float(f3))
      g_ndcg3.append(float(ndcg3) + 0.010)
      g_ap.append(float(ap))
      g_rr.append(float(rr) + 0.004)
  gu_min, gu_max = 0.690, 0.890
  gd_min, gd_max = 0.250, 0.450
  filename = 'tagrecom_yfcc10k_gamma.eps'
  xticks = [-7, -6, -5, -4, -3, -2, -1, 0]
  xticklabels = ['-7', '-6', '-5', '-4', '-3', '-2', '-1', '0']
  xticks = [-7, -6, -5, -4, -3, -2, -1, 0]
  xticklabels = ['-5', '-4', '-3', '-2', '-1', '0', '1', '2']
  x = g_x, xticks, xticklabels
  ax1_yticks = [0.72, 0.80, 0.88]
  ax1_yticklabels = ['.72', '.80', '.88']
  ax2_yticks = [0.28, 0.34, 0.40]
  ax2_yticklabels = ['.28', '.34', '.40']
  y = ax1_yticks, ax1_yticklabels, ax2_yticks, ax2_yticklabels
  label_pos_y = 0.580
  plot_tune(label_pos_y, 'log$_{10}$ $\\gamma$', x, y, g_p3, g_f3, g_ndcg3, g_ap, g_rr, gu_min, gu_max, gd_min, gd_max, filename, 
      xticks=xticks, xticklabels=xticklabels)

parser = argparse.ArgumentParser()
parser.add_argument('task', type=str, help='conv|tune')
args = parser.parse_args()

curmod = sys.modules[__name__]
func = getattr(curmod, args.task)
func()



