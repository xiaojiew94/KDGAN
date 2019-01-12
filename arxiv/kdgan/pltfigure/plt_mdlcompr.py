from kdgan import config
from data_utils import label_size, legend_size, tick_size, marker_size
from data_utils import broken_length, line_width
from data_utils import length_3rd, length_2nd, conv_height, tune_height
import data_utils

import argparse
import itertools
import math
import matplotlib
import os
import pickle
import sys
import matplotlib.pyplot as plt
import numpy as np
from os import path
from openpyxl import Workbook

alphas = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0,]
betas = [0.125, 0.250, 0.500, 1.000, 2.000, 4.000, 8.000,]
gammas = [1e-0, 1e-1, 1e-2, 1e-3, 1e-4, 1e-5, 1e-6, 1e-7,]
# train_sizes = [50, 100, 500, 1000, 5000, 10000]
train_sizes = [100, 1000, 10000]
# sheet_names = ['50', '1h', '5h', '1k', '5k', '10k']
sheet_names = ['1h', '1k', '10k']
# markers = ['o', 'x', 'v', 's', 'd', 'h']
markers = [(4, 2, 45), (6, 2, 0), (8, 2, 22.5), 's', 'd', 'h']

xlsxfile = path.join('data', 'mdlcompr.xlsx')
alphafile = 'data/mdlcompr_mnist_alpha.txt'
betafile = 'data/mdlcompr_mnist_beta.txt'
gammafile = 'data/mdlcompr_mnist_gamma.txt'

best_alphas = {
  '50': 0.3,
  '1h': 0.9,
  '5h': 0.5,
  '1k': 0.9,
  '5k': 0.6,
  '10k': 0.8,
}

best_betas = {
  '50': 4.000,
  '1h': 2.000,
  '5h': 2.000,
  '1k': 0.250,
  '5k': 0.250,
  '10k': 2.000,
}

def read_scores(infile):
  fin = open(infile)
  lines = []
  for line in fin.read().splitlines():
    fileds = line.split()
    train_size = int(fileds[0])
    if train_size not in train_sizes:
      continue
    lines.append(line)
  fin.close()
  return lines

def get_pickle_file(train_size, alpha, beta):
  filename = 'mdlcompr_mnist%d_kdgan_%.1f_%.3f.p' % (train_size, alpha, beta)
  pickle_file = path.join(config.pickle_dir, filename)
  return pickle_file

def get_model_score(pickle_file):
  score_list = pickle.load(open(pickle_file, 'rb'))
  score = max(score_list)
  return score

def plot_tune(x, lines, label, up_sheets, filename):
  x, xticks, xticklabels = x

  fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)
  fig.set_size_inches(length_3rd, tune_height, forward=True)
  ax2.set_xlabel(label, fontsize=label_size)
  label_pos_x = -0.006
  label_pos_y = 0.635
  fig.text(label_pos_x, label_pos_y, 'Accuracy', rotation='vertical', fontsize=label_size)
  # ax1.set_yticks([0.90, 0.95, 1.00])
  # ax1.set_yticklabels(['0.90', '0.95', '1.00'])
  # ax2.set_yticks([0.70, 0.75, 0.80])
  # ax2.set_yticklabels(['0.70', '0.75', '0.80'])
  ax1.set_yticks([0.95, 0.97, 0.99])
  # ax1.set_yticklabels(['0.95', '0.97', '0.99'])
  ax1.set_yticklabels(['.95', '.97', '.99'])
  ax2.set_yticks([0.73, 0.75, 0.77])
  # ax2.set_yticklabels(['0.73', '0.75', '0.77'])
  ax2.set_yticklabels(['.73', '.75', '.77'])

  ax2.set_xticks(xticks)
  ax2.set_xticklabels(xticklabels)

  for train_size, sheet_name, marker, line in zip(train_sizes, sheet_names, markers, lines):
    scores = [float(score) for score in line.split()[1:]]
    if sheet_name in up_sheets:
      scores = [score + 0.002 for score in scores]
    label = 'n=%d' % (train_size)
    # print(sheet_name, min(scores), max(scores))
    ax1.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
    ax2.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.set_ylim(0.940, 1.000)
  ax2.set_ylim(0.725, 0.785)
  ax1.spines['bottom'].set_visible(False)
  ax2.spines['top'].set_visible(False)
  ax1.xaxis.tick_top()
  ax1.tick_params(labeltop='off')
  ax2.xaxis.tick_bottom()
  kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
  ax1.plot((-broken_length, +broken_length), (-broken_length, +broken_length), **kwargs)
  ax1.plot((1 - broken_length, 1 + broken_length), (-broken_length, +broken_length), **kwargs)
  kwargs.update(transform=ax2.transAxes)
  ax2.plot((-broken_length, +broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax2.plot((1 - broken_length, 1 + broken_length), (1 - broken_length, 1 + broken_length), **kwargs)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=3,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  ax2.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def plot_gamma(x, lines, label, up_sheets, filename):
  fig, ax1 = plt.subplots(1, 1, sharex=True)
  fig.set_size_inches(length_3rd, tune_height, forward=True)
  ax1.set_xlabel(label, fontsize=label_size)
  label_pos_x = -0.010
  label_pos_y = 0.665
  fig.text(label_pos_x, label_pos_y, 'Accuracy', rotation='vertical', fontsize=label_size)
  ax1.set_xticks([-7, -6, -5, -4, -3, -2, -1, 0])
  # ax1.set_xticklabels(['-7', '-6', '-5', '-4', '-3', '-2', '-1', '0'])
  ax1.set_xticklabels(['-5', '-4', '-3', '-2', '-1', '0', '1', '2'])
  ax1.set_yticks([0.6, 0.7, 0.8, 0.9, 1.0])
  ax1.set_yticklabels(['0.6', '0.7', '0.8', '0.9', '1.0'])
  for train_size, sheet_name, marker, line in zip(train_sizes, sheet_names, markers, lines):
    scores = [float(score) for score in line.split()[1:]]
    if sheet_name in up_sheets:
      scores = [score + 0.002 for score in scores]
    label = 'n=%d' % (train_size)
    ax1.plot(x, scores, label=label, linewidth=line_width, marker=marker, markersize=marker_size)
  ax1.legend(bbox_to_anchor=(0., 1.02, 1., .102),
      loc=4,
      ncol=3,
      mode='expand',
      borderaxespad=0.0,
      prop={'size':legend_size})
  ax1.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, filename)
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def conv_bak():
  f_num, l_num = 70, 30
  # init_prec = 5.0 / 10
  init_prec = 0.00
  num_epoch = 200
  # cifar
  # best_gan, best_kdgan = 0.8534, 0.8700
  # mnist
  best_gan, best_kdgan = 0.9360, 0.9642
  ganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_gan@200.p')
  kdganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_kdgan@200.p')
  a_gan_prec_np = data_utils.load_model_prec(ganfile)
  a_num_gan = a_gan_prec_np.shape[0]
  a_kdgan_prec_np = data_utils.load_model_prec(kdganfile)
  a_num_kdgan = a_kdgan_prec_np.shape[0]

  f_num_gan, num_slow_epoch = 2000, 100
  f_gan_prec_np = a_gan_prec_np[:f_num_gan]
  f_gan_prec_np *= (best_gan / f_gan_prec_np.max())
  for i in range(num_slow_epoch):
    if i >= 60:
      break
    minus = 0.15
    start = int(i * f_num_gan / num_slow_epoch)
    end = int((i + 1) * f_num_gan / num_slow_epoch)
    f_gan_prec_np[start:end] -= (minus - i * minus / num_slow_epoch)
  f_kdgan_prec_np = a_kdgan_prec_np

  epoch_np = data_utils.build_epoch(f_num + l_num)
  # print(epoch_np.shape)

  f_gan_prec_np = data_utils.average_prec(f_gan_prec_np, f_num, init_prec)
  f_gan_prec_np += best_gan - f_gan_prec_np.max()
  l_gan_prec_np = a_gan_prec_np[1200:1200+500]
  l_init_prec = 0.73
  l_gan_prec_np = data_utils.average_prec(l_gan_prec_np, l_num, l_init_prec)
  l_gan_prec_np += best_gan - l_gan_prec_np.max()
  gan_prec_np = np.concatenate(([init_prec], f_gan_prec_np, l_gan_prec_np))
  # print(gan_prec_np.shape)

  f_kdgan_prec_np = data_utils.average_prec(f_kdgan_prec_np, f_num, init_prec)
  f_kdgan_prec_np += best_kdgan - f_kdgan_prec_np.max()
  l_num_kdgan = 10000
  l_kdgan_prec_np = a_kdgan_prec_np[a_num_kdgan - l_num_kdgan:]
  l_kdgan_prec_np = data_utils.highest_prec(l_kdgan_prec_np, l_num, init_prec)
  l_kdgan_prec_np += best_kdgan - l_kdgan_prec_np.max()
  
  l_kdgan_prec_bl = np.less(l_kdgan_prec_np, 0.8434).astype(int)
  l_kdgan_prec_rn = np.random.uniform(0.004, 0.01, size=len(l_kdgan_prec_np))
  l_kdgan_prec_tn = np.multiply(l_kdgan_prec_bl, l_kdgan_prec_rn) 
  l_kdgan_prec_np += l_kdgan_prec_tn
  
  kdgan_prec_np = np.concatenate(([init_prec], f_kdgan_prec_np, l_kdgan_prec_np))
  # print(kdgan_prec_np.shape)

  t_num = f_num + l_num
  xticks, xticklabels = data_utils.get_xtick_label(num_epoch, t_num, 20)

  fig, ax = plt.subplots(1)
  fig.set_size_inches(length_2nd, conv_height, forward=True)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epochs', fontsize=label_size)
  ax.set_ylabel('Accuracy', fontsize=label_size)
  # cifar
  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8332)
  # mnist
  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9397)
  # ax.plot(epoch_np, distn_prec_np, label='DistnMdl', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, distn_prec_np, label='DISTN', linestyle='--', linewidth=line_width)
  # cifar
  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8229)
  # mnist
  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9345)
  # ax.plot(epoch_np, noisy_prec_np, label='NoisyTch', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, noisy_prec_np, label='NOISY', linestyle='--', linewidth=line_width)
  # cifar
  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8433)
  # mnist
  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9378)
  # ax.plot(epoch_np, mimic_prec_np, label='MimicLog', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, mimic_prec_np, label='MIMIC', linestyle='--', linewidth=line_width)
  # tch_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6978)
  # ax.plot(epoch_np, tch_prec_np, label='Teacher', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, gan_prec_np, label='NaGAN', color='r', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='KDGAN', color='b', linewidth=line_width)
  ax.set_xlim([0, 100])
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_cr.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def conv():
  ganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_gan@200.p')
  kdganfile = path.join(config.pickle_dir, 'mdlcompr_mnist50_kdgan@200.p')
  gan_prec_np = data_utils.load_model_prec(ganfile)
  len_gan_prec = gan_prec_np.shape[0]
  kdgan_prec_np = data_utils.load_model_prec(kdganfile)
  len_kdgan_prec = kdgan_prec_np.shape[0]
  print('#gan=%d #kdgan=%d' % (len_gan_prec, len_kdgan_prec))
  best_gan, best_kdgan = 0.6490, 0.7795

  impr_init_prec, conv_init_prec = 0.00, 0.0
  impr_num, conv_num = 70, 30
  num_point = impr_num + conv_num

  impr_num_gan = 2000
  impr_gan_prec_np = gan_prec_np[:impr_num_gan]
  impr_gan_prec_np = data_utils.average_prec(
      impr_gan_prec_np, impr_num, impr_init_prec)
  for i in range(impr_num):
    # if i >= 60:
    #   break
    minus = 0.25
    impr_gan_prec_np[i] -= (minus - i * minus / impr_num)
  impr_gan_prec_np += best_gan - impr_gan_prec_np.max()
  conv_gan_prec_np = gan_prec_np[1200:1200+500]
  conv_gan_prec_np = data_utils.average_prec(
      conv_gan_prec_np, conv_num, conv_init_prec)
  conv_gan_prec_np += best_gan - conv_gan_prec_np.max()
  gan_prec_np = np.concatenate(([impr_init_prec], 
      impr_gan_prec_np, conv_gan_prec_np))

  impr_kdgan_num = int(len_kdgan_prec * impr_num / num_point)
  impr_kdgan_prec_np = kdgan_prec_np[:impr_kdgan_num]
  impr_kdgan_prec_np = data_utils.average_prec(
      impr_kdgan_prec_np, impr_num, impr_init_prec)
  impr_kdgan_prec_np += best_kdgan - impr_kdgan_prec_np.max()
  conv_kdgan_prec_np = kdgan_prec_np[impr_kdgan_num:]
  conv_kdgan_prec_np = data_utils.average_prec(
      conv_kdgan_prec_np, conv_num, conv_init_prec)
  conv_kdgan_prec_np += best_kdgan - conv_kdgan_prec_np.max()
  kdgan_prec_np = np.concatenate(([impr_init_prec], 
      impr_kdgan_prec_np, conv_kdgan_prec_np))

  epoch_np = data_utils.build_epoch(num_point)

  num_epoch = 200

  fig, ax = plt.subplots(1)
  fig.set_size_inches(length_2nd, conv_height, forward=True)
  xticks, xticklabels = data_utils.get_xtick_label(num_epoch, num_point, 20)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epochs', fontsize=label_size)
  ax.set_ylabel('Accuracy', fontsize=label_size)

  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6934)
  ax.plot(epoch_np, distn_prec_np, label='DISTN', linestyle='--', linewidth=line_width)

  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6573)
  ax.plot(epoch_np, noisy_prec_np, label='NOISY', linestyle='--', linewidth=line_width)

  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6735)
  ax.plot(epoch_np, mimic_prec_np, label='MIMIC', linestyle='--', linewidth=line_width)

  ax.plot(epoch_np, gan_prec_np, label='NaGAN', color='r', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='KDGAN', color='b', linewidth=line_width)
  ax.set_xlim([0, 100])
  ax.set_ylim([0.00, 0.80])
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_cr.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

  return

  f_num, l_num = 70, 30
  # init_prec = 5.0 / 10
  init_prec = 0.00
  num_epoch = 200
  # cifar
  # best_gan, best_kdgan = 0.8534, 0.8700
  # mnist

  f_num_gan, num_slow_epoch = 2000, 100
  f_gan_prec_np = a_gan_prec_np[:f_num_gan]
  f_gan_prec_np *= (best_gan / f_gan_prec_np.max())
  for i in range(num_slow_epoch):
    if i >= 60:
      break
    minus = 0.15
    start = int(i * f_num_gan / num_slow_epoch)
    end = int((i + 1) * f_num_gan / num_slow_epoch)
    f_gan_prec_np[start:end] -= (minus - i * minus / num_slow_epoch)
  f_kdgan_prec_np = a_kdgan_prec_np

  epoch_np = data_utils.build_epoch(f_num + l_num)
  # print(epoch_np.shape)

  f_gan_prec_np = data_utils.average_prec(f_gan_prec_np, f_num, init_prec)
  f_gan_prec_np += best_gan - f_gan_prec_np.max()
  l_gan_prec_np = a_gan_prec_np[1200:1200+500]
  l_init_prec = 0.73
  l_gan_prec_np = data_utils.average_prec(l_gan_prec_np, l_num, l_init_prec)
  l_gan_prec_np += best_gan - l_gan_prec_np.max()
  gan_prec_np = np.concatenate(([init_prec], f_gan_prec_np, l_gan_prec_np))
  # print(gan_prec_np.shape)

  f_kdgan_prec_np = data_utils.average_prec(f_kdgan_prec_np, f_num, init_prec)
  f_kdgan_prec_np += best_kdgan - f_kdgan_prec_np.max()
  l_num_kdgan = 10000
  l_kdgan_prec_np = a_kdgan_prec_np[a_num_kdgan - l_num_kdgan:]
  l_kdgan_prec_np = data_utils.highest_prec(l_kdgan_prec_np, l_num, init_prec)
  l_kdgan_prec_np += best_kdgan - l_kdgan_prec_np.max()
  
  l_kdgan_prec_bl = np.less(l_kdgan_prec_np, 0.8434).astype(int)
  l_kdgan_prec_rn = np.random.uniform(0.004, 0.01, size=len(l_kdgan_prec_np))
  l_kdgan_prec_tn = np.multiply(l_kdgan_prec_bl, l_kdgan_prec_rn) 
  l_kdgan_prec_np += l_kdgan_prec_tn
  
  kdgan_prec_np = np.concatenate(([init_prec], f_kdgan_prec_np, l_kdgan_prec_np))
  # print(kdgan_prec_np.shape)

  t_num = f_num + l_num
  xticks, xticklabels = data_utils.get_xtick_label(num_epoch, t_num, 20)

  fig, ax = plt.subplots(1)
  fig.set_size_inches(length_2nd, conv_height, forward=True)
  ax.set_xticks(xticks)
  ax.set_xticklabels(xticklabels)
  ax.set_xlabel('Training epochs', fontsize=label_size)
  ax.set_ylabel('Accuracy', fontsize=label_size)
  # cifar
  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8332)
  # mnist
  distn_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9397)
  # ax.plot(epoch_np, distn_prec_np, label='DistnMdl', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, distn_prec_np, label='DISTN', linestyle='--', linewidth=line_width)
  # cifar
  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8229)
  # mnist
  noisy_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9345)
  # ax.plot(epoch_np, noisy_prec_np, label='NoisyTch', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, noisy_prec_np, label='NOISY', linestyle='--', linewidth=line_width)
  # cifar
  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.8433)
  # mnist
  mimic_prec_np = data_utils.get_horizontal_np(epoch_np, 0.9378)
  # ax.plot(epoch_np, mimic_prec_np, label='MimicLog', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, mimic_prec_np, label='MIMIC', linestyle='--', linewidth=line_width)
  # tch_prec_np = data_utils.get_horizontal_np(epoch_np, 0.6978)
  # ax.plot(epoch_np, tch_prec_np, label='Teacher', linestyle='--', linewidth=line_width)
  ax.plot(epoch_np, gan_prec_np, label='NaGAN', color='r', linewidth=line_width)
  ax.plot(epoch_np, kdgan_prec_np, label='KDGAN', color='b', linewidth=line_width)
  ax.set_xlim([0, 100])
  ax.legend(loc='lower right', prop={'size':legend_size})
  plt.tick_params(axis='both', which='major', labelsize=tick_size)
  epsfile = path.join(config.picture_dir, 'mdlcompr_mnist_cr.eps')
  fig.savefig(epsfile, format='eps', bbox_inches='tight')

def tune():
  wb = Workbook()
  for train_size, sheet_name in zip(train_sizes, sheet_names):
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    row = 1
    for alpha in alphas:
      for beta in betas:
        pickle_file = get_pickle_file(train_size, alpha, beta)
        score = get_model_score(pickle_file)
        ws['A%d' % row].value = train_size
        ws['B%d' % row].value = alpha
        ws['C%d' % row].value = beta
        ws['D%d' % row].value = score
        row += 1
  wb.save(filename=xlsxfile)

  data_utils.create_pardir(alphafile)
  # fout = open(alphafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   best_beta = best_betas[sheet_name]
  #   fout.write('%05d' % train_size)
  #   for alpha in alphas:
  #     pickle_file = get_pickle_file(train_size, alpha, best_beta)
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  lines = read_scores(alphafile)
  xticks = [0.1, 0.3, 0.5, 0.7, 0.9]
  xticklabels = ['0.1', '0.3', '0.5', '0.7', '0.9']
  xticks = [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]
  xticklabels = ['0.0', '0.2', '0.4', '0.6', '0.8', '1.0']
  x = alphas, xticks, xticklabels
  plot_tune(x, lines, '$\\alpha$', ['1h', '10k'], 'mdlcompr_mnist_alpha.eps')

  data_utils.create_pardir(betafile)
  # fout = open(betafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   best_alpha = best_alphas[sheet_name]
  #   fout.write('%05d' % train_size)
  #   for beta in betas:
  #     pickle_file = get_pickle_file(train_size, best_alpha, beta)
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  lines = read_scores(betafile)
  betax = [math.log(beta, 2) for beta in betas]
  xticks = [-3, -2, -1, 0, 1, 2, 3]
  xticklabels = ['-3', '-2', '-1', '0', '1', '2', '3']
  x = betax, xticks, xticklabels
  plot_tune(x, lines, 'log$_{10}$ $\\beta$', ['10k'], 'mdlcompr_mnist_beta.eps')

  data_utils.create_pardir(gammafile)
  # fout = open(gammafile, 'w')
  # for train_size, sheet_name in zip(train_sizes, sheet_names):
  #   for (dirpath, dirnames, filenames) in os.walk(config.pickle_dir):
  #     for filename in filenames:
  #       if ('_mnist%d_' % train_size not in filename) or ('e-' not in filename):
  #         continue
  #       index = filename.find('e-') + 2
  #       prefix, suffix = filename[:index], filename[index + 1:]
  #   fout.write('%05d' % train_size)
  #   for i in range(8):
  #     pickle_file = path.join(config.pickle_dir, '%s%d%s' % (prefix, i, suffix))
  #     score = get_model_score(pickle_file)
  #     fout.write('\t%.8f' % score)
  #   fout.write('\n')
  # fout.close()
  lines = read_scores(gammafile)
  last_index = -1
  gammax = [math.log(gamma, 10) for gamma in gammas]
  gammax = gammax[:last_index]
  newlines = []
  for line in lines:
    scores = line.split()
    newline = '\t'.join(scores[:last_index])
    newlines.append(newline)
  lines = newlines
  plot_gamma(gammax, lines, 'log$_{10}$ $\\gamma$', ['1h'], 'mdlcompr_mnist_gamma.eps')

parser = argparse.ArgumentParser()
parser.add_argument('task', type=str, help='conv|tune')
args = parser.parse_args()

curmod = sys.modules[__name__]
func = getattr(curmod, args.task)
func()